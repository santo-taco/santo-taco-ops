#!/usr/bin/env python3
"""
================================================================================
  SANTO TACO — COGS DEEP DIVE
================================================================================

  Fetches R365 report emails from Gmail + pulls live sales from Toast,
  and writes two tabs to a dedicated Google Sheet:

    Tab 1 — "Weekly Data"
      Slide-ready numbers for managers to paste into the weekly dashboard deck.
      Divided by location (SoHo then Union Square).
      Sections:
        • Slide 3  — Lunch/Dinner orders, Sales, Discounts  (Toast)
        • Slide 4  — Food actual vs theoretical by taco group   (R365 AVT)
        • Slide 5  — Beverage actual vs theoretical             (R365 AVT)

    Tab 2 — "COGS Deep Dive"
      Ingredients as rows, weeks as columns (appends each run).
      Tracks: ending inventory qty, cost per unit, actual usage, theoretical
      usage, efficiency %, and total $ spend — per ingredient per location.

  Gmail subject lines (must match exactly):
      AVT_SUBJECT_SOHO = "Actual vs Theoretical Analysis Subscription Soho"
      AVT_SUBJECT_USQ  = "Actual vs Theoretical Analysis Subscription Union Square"
      STOCK_SUBJECT    = "R365 Reports: Inventory Stock Count Detail"

  SETUP — add to .env:
      COGS_SHEET_ID=...               (create a blank sheet, paste its ID here)
      TOAST_CLIENT_ID=...
      TOAST_CLIENT_SECRET=...
      SHIFTS_ACCESS_TOKEN=...         (or SHIFTS_API_KEY)
      SHIFTS_COMPANY_ID=...
      SHIFTS_SOHO_LOCATION_ID=...
      SHIFTS_UNIONSQ_LOCATION_ID=...

  Run:
      python cogs_deep_dive.py                          # current week
      python cogs_deep_dive.py --week-ending 2026-03-23 # specific week

================================================================================
"""

import argparse
import base64
import io
import os
import re
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import gspread
from zoneinfo import ZoneInfo

# ── Environment ──────────────────────────────────────────────────────────────────────────────

load_dotenv()

SHARE_WITH_EMAIL      = "robyn@eatsantotaco.com"
GOOGLE_CREDS_FILE     = os.path.join(os.path.dirname(__file__), "google_credentials.json")
SERVICE_ACCOUNT_EMAIL = "santo-taco-sheets@santo-taco-490100.iam.gserviceaccount.com"
NYC                   = ZoneInfo("America/New_York")

COGS_SHEET_ID         = os.getenv("COGS_SHEET_ID", "")
TOAST_CLIENT_ID       = os.getenv("TOAST_CLIENT_ID")
TOAST_CLIENT_SECRET   = os.getenv("TOAST_CLIENT_SECRET")
GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY", "")

TOAST_AUTH_URL   = "https://ws-api.toasttab.com/authentication/v1/authentication/login"
TOAST_ORDERS_URL = "https://ws-api.toasttab.com/orders/v2/ordersBulk"
PLACES_API_URL   = "https://places.googleapis.com/v1/places/{place_id}"

# Toast GUIDs — same as santo_taco_labor_check.py
TOAST_GUIDS = {
    "soho": "fb3ed56a-e127-4fd9-addb-7d71af425727",
    "usq":  "863ea8dc-996c-4e0b-bc7b-274379652680",
}

# Google Places IDs (used to fetch live opening hours)
PLACE_IDS = {
    "soho": "ChIJg9XtpY1ZwokR8hMIfPF27m4",
    "usq":  "ChIJqadlawBZwokRhpBQYJ01mB4",
}

AM_PM_CUTOFF = 16.5  # orders before 4:30 pm = Lunch, at/after = Dinner (matches Toast service periods)

# Dining option GUIDs to exclude from order counts and sales totals.
# Covers both locations — GUIDs are unique per restaurant so safe to combine.
EXCLUDED_DINING_OPTION_GUIDS = {
    # USQ
    "546e422e-d90b-4e7f-9715-392d75b8b23e",  # Full Buyout
    "32f5eaeb-44c9-4aa0-8ef9-dbfd6c54184e",  # Partial Buyout
    "1508b95c-f467-44fd-b5f7-7d268163904f",  # Offsite Event
    "22d09e47-d026-40fb-8cfc-d77f4b77e136",  # Catering Delivery
    "a2ff5954-a1a1-4b5f-b86f-812318947616",  # Catering Pick-Up
    "3a757bb2-d23b-47bb-a83d-d8314ef11036",  # E-Gift Cards
    # SoHo
    "6b5eb0cc-6dc5-49a2-81dd-65a42a16efeb",  # Catering Delivery
    "804dfe94-8355-4577-a31d-6c3b0c73540e",  # Catering Pick-Up
    "b48c42ea-a8f0-48f9-a6aa-0f9802e06a9e",  # E-Gift Cards
}

# AVT ingredient names to exclude from COGS totals (case-insensitive match).
# These appear in both SoHo and USQ AVTs but are not food/bev COGS.
AVT_EXCLUDED_INGREDIENTS: set[str] = {
    "paper goods",
    "kitchen disposables",
    "r&d item",
}

# ── Google Places helpers ─────────────────────────────────────────────────────────────

def fetch_open_hours(place_id: str, start_date: date, end_date: date) -> tuple[float, float]:
    """
    Fetch regularOpeningHours from the Google Places API and return
    (am_hours, pm_hours) for the given week. Lunch = open → 4:30 pm, Dinner = 4:30 pm → close.
    Returns (0.0, 0.0) on any error so callers can fall back gracefully.
    """
    if not GOOGLE_PLACES_API_KEY:
        return 0.0, 0.0
    try:
        resp = requests.get(
            PLACES_API_URL.format(place_id=place_id),
            headers={
                "X-Goog-Api-Key":   GOOGLE_PLACES_API_KEY,
                "X-Goog-FieldMask": "regularOpeningHours",
            },
            timeout=10,
        )
        resp.raise_for_status()
        periods = resp.json().get("regularOpeningHours", {}).get("periods", [])
    except Exception as exc:
        print(f"  [warn] Places API error: {exc} — falling back to order timestamps")
        return 0.0, 0.0

    # Build lookup: Places day (0=Sun, 1=Mon…) → (open_decimal, close_decimal)
    hours_by_day: dict[int, tuple[float, float]] = {}
    for p in periods:
        o = p.get("open", {})
        c = p.get("close", {})
        day       = o.get("day")
        open_dec  = o.get("hour", 0) + o.get("minute", 0) / 60
        close_dec = c.get("hour", 0) + c.get("minute", 0) / 60
        if day is not None:
            hours_by_day[day] = (open_dec, close_dec)

    am_hours = 0.0
    pm_hours = 0.0
    d = start_date
    while d <= end_date:
        places_day = (d.weekday() + 1) % 7   # Python Mon=0 → Places Mon=1, Sun=0
        if places_day in hours_by_day:
            open_dec, close_dec = hours_by_day[places_day]
            if open_dec < AM_PM_CUTOFF:
                am_hours += min(close_dec, AM_PM_CUTOFF) - open_dec
            if close_dec > AM_PM_CUTOFF:
                pm_hours += close_dec - max(open_dec, AM_PM_CUTOFF)
        d += timedelta(days=1)

    return round(am_hours, 1), round(pm_hours, 1)


# ── Toast helpers ────────────────────────────────────────────────────────────────────

def toast_authenticate() -> str:
    resp = requests.post(
        TOAST_AUTH_URL,
        json={
            "clientId":       TOAST_CLIENT_ID,
            "clientSecret":   TOAST_CLIENT_SECRET,
            "userAccessType": "TOAST_MACHINE_CLIENT",
        },
        timeout=30,
    )
    resp.raise_for_status()
    data  = resp.json()
    token = (data.get("token") or {}).get("accessToken")
    if not token:
        sys.exit(f"[ERROR] Unexpected Toast auth response: {data}")
    return token


def _parse_next_link(link_header: str | None) -> str | None:
    if not link_header:
        return None
    for part in link_header.split(","):
        part = part.strip()
        if 'rel="next"' in part:
            url_part = part.split(";")[0].strip()
            return url_part.lstrip("<").rstrip(">")
    return None


def toast_fetch_sales_and_orders(
    token:           str,
    restaurant_guid: str,
    start_date:      date,
    end_date:        date,
) -> dict:
    """
    Fetch all orders for the week and return:
    {
        "net_sales":  float,
        "am_orders":  int,   # orders opened before 16:00 NYC
        "pm_orders":  int,   # orders opened at or after 16:00 NYC
    }
    AM/PM split is by the local NYC open time of each order.
    """
    headers = {
        "Authorization":                f"Bearer {token}",
        "Toast-Restaurant-External-ID": restaurant_guid,
    }

    next_url: str | None = TOAST_ORDERS_URL
    params: dict | None  = {
        "startDate": start_date.strftime("%Y-%m-%dT00:00:00.000+0000"),
        "endDate":   (end_date + timedelta(days=1)).strftime("%Y-%m-%dT23:59:59.000+0000"),
        "pageSize":  100,
    }

    all_orders: list[dict] = []
    page = 1
    while next_url:
        resp = requests.get(next_url, headers=headers, params=params, timeout=60)
        resp.raise_for_status()
        batch = resp.json()
        all_orders.extend(batch)
        print(f"          page {page}: {len(batch)} orders", flush=True)
        next_url = _parse_next_link(resp.headers.get("link"))
        params   = None
        page    += 1
        if not batch:
            break

    net_sales    = 0.0
    am_orders    = 0
    pm_orders    = 0
    total_orders = 0   # all orders incl. voided — matches Toast Service Mode Summary
    discounts: dict[str, float] = defaultdict(float)  # discount name → total $ amount
    # Per-day first/last order time, split by Lunch (<16:30) and Dinner (≥16:30)
    daily_times: dict[date, dict[str, list]] = defaultdict(lambda: {"am": [], "pm": []})

    for order in all_orders:
        # Filter to requested week by businessDate first (applies to all orders)
        bd_raw = order.get("businessDate")
        if not bd_raw:
            continue
        try:
            s  = str(bd_raw)
            bd = date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except (ValueError, TypeError, IndexError):
            continue
        if not (start_date <= bd <= end_date):
            continue

        # Skip voided orders for all other calculations
        if order.get("voided"):
            continue

        # AM/PM split by openedDate local time
        opened_raw = order.get("openedDate")
        is_am      = True
        opened_nyc = None
        if opened_raw:
            try:
                opened_utc = datetime.fromisoformat(
                    str(opened_raw).replace("Z", "+00:00")
                )
                opened_nyc = opened_utc.astimezone(NYC)
                dec_hour   = opened_nyc.hour + opened_nyc.minute / 60
                is_am = dec_hour < AM_PM_CUTOFF   # before 4:30pm = Lunch
            except (ValueError, TypeError):
                pass

        # Count all non-voided orders — matches Toast/P&L total order count
        # (includes catering/buyout, excludes only voided orders)
        total_orders += 1

        # Exclude buyouts, catering, gift cards from AM/PM counts and net_sales
        do_guid = (order.get("diningOption") or {}).get("guid")
        if do_guid in EXCLUDED_DINING_OPTION_GUIDS:
            continue

        # Track opened time for open-hours calculation (non-excluded orders)
        if opened_nyc:
            bucket = "am" if is_am else "pm"
            daily_times[bd][bucket].append(opened_nyc)

        order_has_sales = False
        for check in order.get("checks", []):
            if check.get("voided") or check.get("deleted"):
                continue
            gift_card_amt = sum(
                float(sel.get("price") or 0)
                for sel in check.get("selections", [])
                if not sel.get("voided")
                and "gift card" in (
                    sel.get("displayName") or sel.get("name") or ""
                ).lower()
            )
            net = max(float(check.get("amount") or 0) - gift_card_amt, 0.0)
            net_sales += net
            if net > 0:
                order_has_sales = True

            # Collect applied discounts — both check-level and selection-level
            for disc in check.get("appliedDiscounts", []):
                name = disc.get("name") or "Unknown Discount"
                amt  = abs(float(disc.get("discountAmount") or 0))
                if amt > 0:
                    discounts[name] += amt
            for sel in check.get("selections", []):
                if sel.get("voided"):
                    continue
                for disc in sel.get("appliedDiscounts", []):
                    name = disc.get("name") or "Unknown Discount"
                    amt  = abs(float(disc.get("discountAmount") or 0))
                    if amt > 0:
                        discounts[name] += amt

        if order_has_sales:
            if is_am:
                am_orders += 1
            else:
                pm_orders += 1

    # Compute actual service hours: first order → last order per period per day
    am_hours = sum(
        (max(t["am"]) - min(t["am"])).total_seconds() / 3600
        for t in daily_times.values()
        if len(t["am"]) >= 2
    )
    pm_hours = sum(
        (max(t["pm"]) - min(t["pm"])).total_seconds() / 3600
        for t in daily_times.values()
        if len(t["pm"]) >= 2
    )

    return {
        "net_sales":    net_sales,
        "am_orders":    am_orders,
        "pm_orders":    pm_orders,
        "total_orders": total_orders,  # matches Toast UI — used for avg check
        "am_hours":     round(am_hours, 1),
        "pm_hours":     round(pm_hours, 1),
        "discounts":    dict(discounts),  # {name: total_amt} sorted by value desc
    }

AVT_SUBJECT       = "R365 Reports: Actual vs Theoretical Analysis"
AVT_SUBJECT_SOHO  = AVT_SUBJECT   # kept for backwards compat
AVT_SUBJECT_USQ   = AVT_SUBJECT
STOCK_SUBJECT     = "R365 Reports: Inventory Stock Count Detail"

# ── Taco group → sub-ingredient mapping (slide 4) ─────────────────────────────
# SoHo uses subrecipe-level rows (R365 rolls up raw ingredients into a recipe).
# USQ reports raw ingredients directly — its AVT is not configured with the same
# subrecipe layer, likely because USQ also operates as a commissary prep kitchen.
# Efficiency is weighted average by theoretical quantity across sub-ingredients.

FOOD_GROUPS_SOHO: dict[str, list[str]] = {
    "Carnitas":              ["Recipes: Carnitas"],
    "Zucchini Taco":         ["Zucchini"],
    "NY Strip":              ["Subrecipe: NYStripLoin sliced"],
    "Sirloin":               ["Subrecipe: Sirloin sliced"],
    "Pollo Taco":            ["Subrecipe: Pollo clean"],
    "Mushroom Taco":         [
        "Shiitake Mushrooms",
        "Cremini Mushrooms",
    ],
    "Avocado":               ["Avocado Hass"],
    "Tuna Saku":             ["Tuna Saku block frozen"],
}

FOOD_GROUPS_USQ: dict[str, list[str]] = {
    "Carnitas":              [
        "Pork Bellies",
        "pork spare ribs 3# dn (usf",
    ],
    "Zucchini Taco":         ["Zucchini"],
    "NY Strip":              ["Beef Strip Loin #180 0X1 [Choice]"],
    "Sirloin":               ["Beef Flap [Soft] Meat #185A [Choice]"],
    "Pollo Taco":            ["Chicken Thighs Boneless Skinless Fresh"],
    "Shrimp Taco":           ["Shrimp 21/25"],
    "Mushroom Taco":         [
        "Shiitake Mushrooms",
        "Cremini Mushrooms",
    ],
    "Avocado":               ["Avocado Hass"],
    "Tuna Saku":             ["Tuna Saku block frozen"],
}

# ── Beverage items (slide 5) ───────────────────────────────────────────────────
# Listed in display order. Matched case-insensitively against AVT ingredient names.
# USQ has keg products instead of bottled Modelo.

BEVERAGE_ITEMS_SOHO: list[str] = [
    "Santo Taco Water Still",
    "Diet Coke",
    "Mexican Coke",
    "Topo Chico",
]

BEVERAGE_ITEMS_USQ: list[str] = [
    "Modelo Negra Keg",
    "Coronita",
    "Modelito",
    "Modelo Especial Keg",
    "Pacifico Clara Keg",
    "Tequila Dobel",
    "Santo Taco Water Still",
    "Diet Coke",
    "Mexican Coke",
    "Topo Chico",
]

# ── Colors ─────────────────────────────────────────────────────────────────────

MINT_GREEN  = {"red": 0.824, "green": 0.906, "blue": 0.804}   # #D2E7CD
DARK_GREEN  = {"red": 0.643, "green": 0.851, "blue": 0.608}   # #A4D99B
WHITE       = {"red": 1.0,   "green": 1.0,   "blue": 1.0}
LIGHT_GREY  = {"red": 0.9,   "green": 0.9,   "blue": 0.9}
DARK_GREY   = {"red": 0.75,  "green": 0.75,  "blue": 0.75}

# ── Date helpers ───────────────────────────────────────────────────────────────

def most_recent_sunday(today: date | None = None) -> date:
    d = today or date.today()
    days_since_sunday = (d.weekday() + 1) % 7
    return d - timedelta(days=days_since_sunday)


# ── Google Sheets client ───────────────────────────────────────────────────────

def _get_sheets_client() -> gspread.Client:
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
    return gspread.authorize(creds)


def _get_gmail_service():
    scopes = ["https://www.googleapis.com/auth/gmail.readonly"]
    creds = Credentials.from_service_account_file(
        GOOGLE_CREDS_FILE, scopes=scopes, subject=SHARE_WITH_EMAIL
    )
    return build("gmail", "v1", credentials=creds)

# ── Gmail attachment fetcher ───────────────────────────────────────────────────

def fetch_excel_attachment(subject: str) -> bytes | None:
    """
    Search Gmail for the most recent email matching *subject*, download the
    first Excel attachment, and return raw bytes.  Returns None on any failure.
    """
    try:
        svc = _get_gmail_service()
    except Exception as exc:
        print(f"  [warn] Gmail auth failed: {exc}")
        return None

    try:
        results = svc.users().messages().list(
            userId="me",
            q=f'subject:"{subject}"',
            maxResults=1,
        ).execute()
    except Exception as exc:
        print(f"  [warn] Gmail search failed for '{subject}': {exc}")
        return None

    messages = results.get("messages", [])
    if not messages:
        print(f"  [warn] No email found: '{subject}'")
        return None

    msg_id = messages[0]["id"]
    try:
        msg = svc.users().messages().get(
            userId="me", id=msg_id, format="full"
        ).execute()
    except Exception as exc:
        print(f"  [warn] Could not fetch email '{subject}': {exc}")
        return None

    parts = msg.get("payload", {}).get("parts", [])
    for part in parts:
        fname = part.get("filename", "")
        if fname.lower().endswith((".xlsx", ".xls")):
            att_id = part.get("body", {}).get("attachmentId")
            if att_id:
                try:
                    att = svc.users().messages().attachments().get(
                        userId="me", messageId=msg_id, id=att_id
                    ).execute()
                    data = base64.urlsafe_b64decode(att["data"])
                    print(f"  [ok]   Attachment found for '{subject}': {fname}")
                    return data
                except Exception as exc:
                    print(f"  [warn] Attachment download failed: {exc}")
                    return None

    print(f"  [warn] No Excel attachment in email '{subject}'")
    return None


def fetch_avt_attachments(subject: str) -> tuple[bytes | None, bytes | None]:
    """
    Fetch up to 2 emails matching *subject* and return (soho_bytes, usq_bytes).
    SoHo vs USQ is identified by 'soho'/'kenmare' or 'union'/'usq' in the
    attachment filename.  If only one email is found, it is assigned by filename.
    """
    try:
        svc = _get_gmail_service()
    except Exception as exc:
        print(f"  [warn] Gmail auth failed: {exc}")
        return None, None

    try:
        results = svc.users().messages().list(
            userId="me",
            q=f'subject:"{subject}"',
            maxResults=2,
        ).execute()
    except Exception as exc:
        print(f"  [warn] Gmail search failed for '{subject}': {exc}")
        return None, None

    messages = results.get("messages", [])
    if not messages:
        print(f"  [warn] No emails found: '{subject}'")
        return None, None

    soho_bytes = None
    usq_bytes  = None

    for msg_meta in messages:
        try:
            msg = svc.users().messages().get(
                userId="me", id=msg_meta["id"], format="full"
            ).execute()
        except Exception as exc:
            print(f"  [warn] Could not fetch email: {exc}")
            continue

        parts = msg.get("payload", {}).get("parts", [])
        for part in parts:
            fname = part.get("filename", "")
            if not fname.lower().endswith((".xlsx", ".xls")):
                continue
            att_id = part.get("body", {}).get("attachmentId")
            if not att_id:
                continue
            try:
                att  = svc.users().messages().attachments().get(
                    userId="me", messageId=msg_meta["id"], id=att_id
                ).execute()
                data = base64.urlsafe_b64decode(att["data"])
            except Exception as exc:
                print(f"  [warn] Attachment download failed: {exc}")
                continue

            fname_lower = fname.lower()
            if any(k in fname_lower for k in ("soho", "kenmare")):
                print(f"  [ok]   AVT SoHo attachment: {fname}")
                soho_bytes = data
            elif any(k in fname_lower for k in ("union", "usq", "university")):
                print(f"  [ok]   AVT Union Square attachment: {fname}")
                usq_bytes = data
            else:
                # Can't tell from filename — assign to whichever slot is empty
                if soho_bytes is None:
                    print(f"  [ok]   AVT (assigned SoHo): {fname}")
                    soho_bytes = data
                elif usq_bytes is None:
                    print(f"  [ok]   AVT (assigned USQ): {fname}")
                    usq_bytes = data
            break  # one attachment per email

    return soho_bytes, usq_bytes

# ── AVT parser ─────────────────────────────────────────────────────────────────

# Column indices in the AVT Excel (0-based, confirmed from sample files)
_AVT_COL = {
    "item":      6,
    "uom":       7,
    "unit_cost": 8,
    "begin_qty": 10,
    "purch_qty": 11,
    "end_qty":   13,
    "actl_qty":  14,
    "theo_qty":  15,
    "effcy":     20,
    "actl_usd":  26,
    "theo_usd":  27,
}

# Metadata rows (0-based)
_AVT_BEGIN_ROW   = 2   # "Begin Inventory: ..."
_AVT_END_ROW     = 4   # "End Inventory   : ..."
_AVT_SALES_ROW   = 5   # "Net Sales      $XX,XXX.XX"
_AVT_DATA_START  = 9   # First data row (Cogs summary)
_AVT_HEADER_ROW  = 8   # Header row


def _safe_float(val, default: float = 0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def parse_avt(raw: bytes, sheet_index: int = 0) -> dict:
    """
    Parse an Actual vs Theoretical Excel file.

    sheet_index: 0 = first sheet (SoHo), 1 = second sheet (USQ).
    If the file has only one sheet, sheet_index is ignored.

    Returns:
    {
        "begin_date": date,
        "end_date":   date,
        "net_sales":  float,
        "ingredients": {
            item_name_lower: {
                "name": str, "uom": str, "unit_cost": float,
                "begin_qty": float, "purch_qty": float, "end_qty": float,
                "actl_qty": float,  "theo_qty": float,  "effcy": float,
                "actl_usd": float,  "theo_usd": float,
            }
        }
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    idx = min(sheet_index, len(wb.worksheets) - 1)
    ws = wb.worksheets[idx]

    rows = list(ws.iter_rows(values_only=True))

    # ── Metadata ──────────────────────────────────────────────────────────────
    def _extract_date(cell_val: str) -> date | None:
        if not cell_val:
            return None
        # Handle non-breaking spaces (\xa0) in the cell value
        cleaned = str(cell_val).replace("\xa0", " ")
        m = re.search(r"(\d{2}/\d{2}/\d{4})", cleaned)
        if m:
            month, day, year = m.group(1).split("/")
            return date(int(year), int(month), int(day))
        return None

    begin_date = _extract_date(rows[_AVT_BEGIN_ROW][4])
    end_date   = _extract_date(rows[_AVT_END_ROW][4])

    sales_cell = str(rows[_AVT_SALES_ROW][4] or "")
    net_sales  = _safe_float(re.sub(r"[^\d.]", "", sales_cell.split("$")[-1]))

    # ── Ingredient rows ───────────────────────────────────────────────────────
    ingredients: dict[str, dict] = {}
    c = _AVT_COL

    for row in rows[_AVT_DATA_START:]:
        item = row[c["item"]]
        uom  = row[c["uom"]]

        # Ingredient rows have a UOM; skip blanks, subtotals, category headers
        if not item or not uom:
            continue
        item_str = str(item).strip()
        if "Total" in item_str or "total" in item_str:
            continue

        # Skip non-food/bev items that appear in the AVT but are not COGS
        if item_str.lower() in AVT_EXCLUDED_INGREDIENTS:
            continue

        ingredients[item_str.lower()] = {
            "name":      item_str,
            "uom":       str(uom).strip(),
            "unit_cost": _safe_float(row[c["unit_cost"]]),
            "begin_qty": _safe_float(row[c["begin_qty"]]),
            "purch_qty": _safe_float(row[c["purch_qty"]]),
            "end_qty":   _safe_float(row[c["end_qty"]]),
            "actl_qty":  _safe_float(row[c["actl_qty"]]),
            "theo_qty":  _safe_float(row[c["theo_qty"]]),
            "effcy":     _safe_float(row[c["effcy"]]),
            "actl_usd":  _safe_float(row[c["actl_usd"]]),
            "theo_usd":  _safe_float(row[c["theo_usd"]]),
        }

    return {
        "begin_date":  begin_date,
        "end_date":    end_date,
        "net_sales":   net_sales,
        "ingredients": ingredients,
    }

# ── Stock Count parser ─────────────────────────────────────────────────────────

# Column indices in Stock Count Detail Excel (0-based, confirmed from sample)
_SC_COL = {
    "location_num":  0,
    "location_name": 1,
    "count_date":    2,
    "item":          6,
    "gl_account":    7,
    "uom":           9,
    "qty":           10,
    "cost_per_item": 11,
    "total_amount":  12,
}
_SC_DATA_START = 3   # Row 0=title, 1=blank, 2=header, 3=first data row


def parse_stock_count(raw: bytes) -> dict:
    """
    Parse Inventory Stock Count Detail Excel.

    Returns:
    {
        "soho":  { item_name_lower: {"name", "uom", "qty", "cost", "total", "count_date"} },
        "usq":   { item_name_lower: {"name", "uom", "qty", "cost", "total", "count_date"} },
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    result = {"soho": {}, "usq": {}}
    c = _SC_COL

    for row in rows[_SC_DATA_START:]:
        loc_num  = row[c["location_num"]]
        item     = row[c["item"]]
        if not item or not loc_num:
            continue

        item_str = str(item).strip()
        key      = item_str.lower()
        count_dt = row[c["count_date"]]
        if hasattr(count_dt, "date"):
            count_dt = count_dt.date()

        entry = {
            "name":       item_str,
            "uom":        str(row[c["uom"]] or "").strip(),
            "qty":        _safe_float(row[c["qty"]]),
            "cost":       _safe_float(row[c["cost_per_item"]]),
            "total":      _safe_float(row[c["total_amount"]]),
            "count_date": count_dt,
        }

        if str(loc_num).strip() == "1":
            result["soho"][key] = entry
        elif str(loc_num).strip() == "2":
            result["usq"][key] = entry

    return result

# ── Data assembly ──────────────────────────────────────────────────────────────

def build_slide3_block(
    avt:            dict,
    location_label: str,
    toast_data:     dict | None = None,
) -> list[list]:
    """
    Returns rows for the Slide 3 section (P&L summary).
    COGS $ and COGS % now come from the P&L Data tab written by pl_report.py.
    AVT data is used for unit cost trends and efficiency only.
    Net Sales, Lunch/Dinner orders come from Toast when available.
    """
    # Use Toast net sales as both display and COGS% denominator when available
    # — matches the P&L methodology. Falls back to AVT net sales if Toast unavailable.
    avt_sales = avt["net_sales"]
    if toast_data and toast_data.get("net_sales", 0) > 0:
        display_sales = toast_data["net_sales"]
    else:
        display_sales = avt_sales

    cogs_usd = sum(
        ing["actl_usd"]
        for ing in avt["ingredients"].values()
        if ing["actl_usd"] > 0
    )
    # NOTE: Total COGS $ and % are no longer written to Slide 3.
    # These now come from the P&L Data tab written by pl_report.py.
    # AVT data is retained for unit cost trends and efficiency only.

    am       = toast_data["am_orders"]    if toast_data else "N/A"
    pm       = toast_data["pm_orders"]    if toast_data else "N/A"
    total    = toast_data["total_orders"] if toast_data else "N/A"
    am_hours = round(toast_data["am_hours"], 1) if toast_data and toast_data.get("am_hours") else "N/A"
    pm_hours = round(toast_data["pm_hours"], 1) if toast_data and toast_data.get("pm_hours") else "N/A"

    # Discounts: roll up all STORYTIME codes into one line, then sort by amount desc
    disc_rows = []
    if toast_data and toast_data.get("discounts"):
        storytime_total = 0.0
        non_storytime: dict[str, float] = {}
        for name, amt in toast_data["discounts"].items():
            if "storytime" in name.lower():
                storytime_total += amt
            else:
                non_storytime[name] = amt
        if storytime_total > 0:
            non_storytime["STORYTIME"] = storytime_total
        sorted_discs = sorted(non_storytime.items(), key=lambda x: x[1], reverse=True)
        for disc_name, disc_amt in sorted_discs:
            disc_rows.append([f"Discount: {disc_name}", f"${disc_amt:,.2f}", "", ""])

    rows = [
        [f"SLIDE 3 — {location_label}"],
        ["Metric", "This Week", "Last Week", "Difference"],
        ["Lunch Orders",   am,                   "",  ""],
        ["Dinner Orders",  pm,                   "",  ""],
        ["Total Orders",   total,                "",  ""],
        ["Lunch Hours",    am_hours,             "",  ""],
        ["Dinner Hours",   pm_hours,             "",  ""],
        ["Net Sales",      f"${display_sales:,.2f}",     "",  ""],
        ["--- Discounts ---", "", "", ""],
        *disc_rows,
        [],
    ]
    return rows


def _weighted_effcy(ingredient_keys: list[str], ingredients: dict) -> float:
    """
    Weighted average efficiency across sub-ingredients,
    weighted by theoretical quantity.
    """
    theo_total = 0.0
    weighted   = 0.0
    for key in ingredient_keys:
        ing = ingredients.get(key.lower())
        if not ing:
            continue
        t = abs(ing["theo_qty"])
        theo_total += t
        weighted   += abs(ing["effcy"]) * t
    if theo_total == 0:
        return 0.0
    return weighted / theo_total


def build_slide4_block(avt: dict, location_label: str, food_groups: dict | None = None) -> list[list]:
    """
    Returns rows for the Slide 4 section (Food Actual vs Theoretical).
    One row per taco group with sub-ingredient detail rows beneath.
    Efficiency is weighted average by theoretical qty across sub-ingredients.
    food_groups defaults to FOOD_GROUPS_SOHO if not provided.
    """
    ings = avt["ingredients"]
    if food_groups is None:
        food_groups = FOOD_GROUPS_SOHO

    header = [f"SLIDE 4 — {location_label} — Food Actual vs Theoretical"]
    col_hdr = [
        "Taco", "Item", "UOM",
        "Beginning", "Purchases", "Ending",
        "Actual", "Theoretical", "Efficiency %",
    ]

    rows = [header, col_hdr]

    for taco_label, sub_keys in food_groups.items():
        blended_effcy = _weighted_effcy(sub_keys, ings)

        first = True
        for key in sub_keys:
            ing = ings.get(key.lower())
            if not ing:
                # still emit a placeholder row so layout is preserved
                ing = {
                    "name": key, "uom": "", "begin_qty": 0, "purch_qty": 0,
                    "end_qty": 0, "actl_qty": 0, "theo_qty": 0, "effcy": 0,
                }

            row = [
                taco_label if first else "",
                ing["name"],
                ing["uom"],
                _fmt_qty(ing["begin_qty"]),
                _fmt_qty(ing["purch_qty"]),
                _fmt_qty(ing["end_qty"]),
                _fmt_qty(ing["actl_qty"]),
                _fmt_qty(ing["theo_qty"]),
                f"{blended_effcy:.0%}" if first else "",
            ]
            rows.append(row)
            first = False

    rows.append([])
    return rows


def build_slide5_block(avt: dict, location_label: str, beverage_items: list | None = None) -> list[list]:
    """
    Returns rows for the Slide 5 section (Beverage Actual vs Theoretical).
    One row per beverage item.
    beverage_items defaults to BEVERAGE_ITEMS_SOHO if not provided.
    """
    ings = avt["ingredients"]
    if beverage_items is None:
        beverage_items = BEVERAGE_ITEMS_SOHO

    header  = [f"SLIDE 5 — {location_label} — Beverage Actual vs Theoretical"]
    col_hdr = [
        "Item", "UOM",
        "Beginning", "Purchases", "Ending",
        "Actual", "Theoretical", "Efficiency %",
    ]

    rows = [header, col_hdr]

    for bev_name in beverage_items:
        ing = ings.get(bev_name.lower())
        if not ing:
            ing = {
                "name": bev_name, "uom": "", "begin_qty": 0, "purch_qty": 0,
                "end_qty": 0, "actl_qty": 0, "theo_qty": 0, "effcy": 0,
            }
        rows.append([
            ing["name"],
            ing["uom"],
            _fmt_qty(ing["begin_qty"]),
            _fmt_qty(ing["purch_qty"]),
            _fmt_qty(ing["end_qty"]),
            _fmt_qty(ing["actl_qty"]),
            _fmt_qty(ing["theo_qty"]),
            f"{ing['effcy']:.0%}",
        ])

    rows.append([])
    return rows


def _fmt_qty(val: float) -> str:
    if val == 0:
        return "0"
    return f"{val:,.3f}".rstrip("0").rstrip(".")

# ── Weekly Data tab builder ────────────────────────────────────────────────────

def build_weekly_data(
    avt_soho:    dict,
    avt_usq:     dict,
    end_date:    date,
    toast_soho:  dict | None = None,
    toast_usq:   dict | None = None,
) -> list[list]:
    """
    Assemble all rows for the Weekly Data tab.
    Structure: title → SoHo block (slides 3,4,5) → USQ block (slides 3,4,5).
    toast_soho / toast_usq are dicts with keys: net_sales, am_orders, pm_orders.
    """
    title = [[f"Santo Taco — Weekly COGS Data — Week ending {end_date.isoformat()}"], []]

    soho_rows = (
        build_slide3_block(avt_soho, "SoHo", toast_data=toast_soho) +
        build_slide4_block(avt_soho, "SoHo", food_groups=FOOD_GROUPS_SOHO) +
        build_slide5_block(avt_soho, "SoHo", beverage_items=BEVERAGE_ITEMS_SOHO)
    )

    usq_rows = (
        build_slide3_block(avt_usq, "Union Square", toast_data=toast_usq) +
        build_slide4_block(avt_usq, "Union Square", food_groups=FOOD_GROUPS_USQ) +
        build_slide5_block(avt_usq, "Union Square", beverage_items=BEVERAGE_ITEMS_USQ)
    )

    return title + soho_rows + [[]] + usq_rows

# ── COGS Deep Dive tab builder ─────────────────────────────────────────────────

# Metrics tracked per ingredient per week
_DEEP_DIVE_METRICS = [
    "End Qty",
    "Cost / Unit ($)",
    "Actual Qty",
    "Theoretical Qty",
    "Efficiency %",
    "Actual Spend ($)",
]

# Number of fixed left columns before week data starts
_DD_LEFT_COLS = 4   # Location | Item | UOM | Metric

# ── Curated ingredient tracking list ──────────────────────────────────────────
# Grouped by category for readability. Names must match AVT ingredient names
# exactly (case-insensitive match is applied at runtime).
# Items marked USQ-only will show blank for SoHo rows.

_DEEP_DIVE_ITEMS: list[dict] = [
    # ── Proteins ──────────────────────────────────────────────────────────────
    # SoHo uses subrecipe names; USQ reports the same proteins as raw ingredients
    # because R365 subrecipes are not configured at USQ.
    {"name": "Recipes: Carnitas",             "display": "Carnitas",  "category": "Protein",
     "usq_names": ["Pork Bellies", "pork spare ribs 3# dn (usf"]},
    {"name": "Subrecipe: NYStripLoin sliced", "display": "NY Strip",  "category": "Protein",
     "usq_names": ["Beef Strip Loin #180 0X1 [Choice]"]},
    {"name": "Subrecipe: Sirloin sliced",     "display": "Sirloin",   "category": "Protein",
     "usq_names": ["Beef Flap [Soft] Meat #185A [Choice]"]},
    {"name": "Subrecipe: Pollo clean",        "display": "Chicken",   "category": "Protein",
     "usq_names": ["Chicken Thighs Boneless Skinless Fresh"]},
    {"name": "Shrimp 21/25",                  "display": "Shrimp",    "category": "Protein",  "usq_only": True},
    # ── Seafood ───────────────────────────────────────────────────────────────
    {"name": "Tuna Saku block frozen",        "display": "Tuna Saku", "category": "Seafood",
     "usq_names": ["Tuna Saku block frozen"]},
    # ── Produce ───────────────────────────────────────────────────────────────
    {"name": "Zucchini",                      "display": "Zucchini",         "category": "Produce"},
    {"name": "Avocado Hass",                  "display": "Avocado",          "category": "Produce"},
    {"name": "Shiitake Mushrooms",            "display": "Shiitake",         "category": "Produce"},
    {"name": "Cremini Mushrooms",             "display": "Cremini",          "category": "Produce"},
    {"name": "Limes",                         "display": "Limes",            "category": "Produce"},
    {"name": "Key Lime",                      "display": "Key Lime",         "category": "Produce"},
    # ── Dry Goods ─────────────────────────────────────────────────────────────
    {"name": "Graza Frizzle",                "display": "Graza Frizzle",    "category": "Dry Goods"},
    # ── Beverages — Non-Alc ───────────────────────────────────────────────────
    {"name": "Santo Taco Water Still",        "display": "SANTO Water",      "category": "Bev Non-Alc"},
    {"name": "Diet Coke",                     "display": "Diet Coke",        "category": "Bev Non-Alc"},
    {"name": "Mexican Coke",                  "display": "Mexican Coke",     "category": "Bev Non-Alc"},
    {"name": "Topo Chico",                    "display": "Topo Chico",       "category": "Bev Non-Alc"},
    {"name": "Bero Beer",                     "display": "Bero Beer",        "category": "Bev Non-Alc"},
    # ── Beverages — Alc ───────────────────────────────────────────────────────
    {"name": "Coronita",                      "display": "Coronita",         "category": "Bev Alc",  "usq_only": True},
    {"name": "Modelito",                      "display": "Modelito",         "category": "Bev Alc",  "usq_only": True},
    {"name": "Modelo Especial Keg",           "display": "Modelo Especial",  "category": "Bev Alc",  "usq_only": True},
    {"name": "Modelo Negra Keg",              "display": "Modelo Negra",     "category": "Bev Alc",  "usq_only": True},
    {"name": "Pacifico Clara Keg",            "display": "Pacifico Clara",   "category": "Bev Alc",  "usq_only": True},
    {"name": "Tequila Dobel",                "display": "Tequila Dobel",    "category": "Bev Alc",  "usq_only": True},
]

# Build fast lookup: lower(name) → item config
_DEEP_DIVE_LOOKUP = {item["name"].lower(): item for item in _DEEP_DIVE_ITEMS}

# Build display name lookup: lower(name) → display label
_DEEP_DIVE_DISPLAY = {
    item["name"].lower(): item.get("display", item["name"])
    for item in _DEEP_DIVE_ITEMS
}


def _dd_week_col_label(end_date: date) -> str:
    return f"Wk {end_date.isoformat()}"


def build_deep_dive_rows(
    avt_soho:    dict,
    avt_usq:     dict,
    stock_soho:  dict,
    stock_usq:   dict,
    end_date:    date,
) -> list[dict]:
    """
    Returns a list of row dicts for the COGS Deep Dive tab.
    Only includes ingredients in _DEEP_DIVE_ITEMS.
    USQ-only items show blank values for SoHo.
    """
    week_label = _dd_week_col_label(end_date)
    records    = []

    for loc_label, avt, stock in [
        ("SoHo",         avt_soho, stock_soho),
        ("Union Square", avt_usq,  stock_usq),
    ]:
        for item_cfg in _DEEP_DIVE_ITEMS:
            item_name  = item_cfg["name"]
            item_lower = item_name.lower()
            usq_only   = item_cfg.get("usq_only", False)
            usq_names  = item_cfg.get("usq_names", [])  # raw ingredient names at USQ

            # Blank row for SoHo if this is a USQ-only ingredient
            if usq_only and loc_label == "SoHo":
                values = {m: "" for m in _DEEP_DIVE_METRICS}
                ing_usq = avt_usq["ingredients"].get(item_lower, {})
                uom = ing_usq.get("uom", "")

            # USQ protein with mapped raw ingredient names — aggregate them
            elif loc_label == "Union Square" and usq_names:
                ings_matched = [
                    avt["ingredients"][n.lower()]
                    for n in usq_names
                    if n.lower() in avt["ingredients"]
                ]
                if ings_matched:
                    end_qty  = sum(i["end_qty"]  for i in ings_matched)
                    actl_qty = sum(i["actl_qty"] for i in ings_matched)
                    theo_qty = sum(i["theo_qty"] for i in ings_matched)
                    actl_usd = sum(i["actl_usd"] for i in ings_matched)
                    # Weighted average efficiency by theoretical qty
                    effcy = _weighted_effcy(usq_names, avt["ingredients"])
                    # Cost / unit: weighted average by actual qty
                    total_actl = sum(i["actl_qty"] for i in ings_matched)
                    cost_u = (
                        sum(i["unit_cost"] * i["actl_qty"] for i in ings_matched) / total_actl
                        if total_actl else 0.0
                    )
                    uom = ings_matched[0]["uom"]
                else:
                    end_qty = actl_qty = theo_qty = actl_usd = effcy = cost_u = 0.0
                    uom = ""

                values = {
                    "End Qty":          _fmt_qty(end_qty),
                    "Cost / Unit ($)":  f"{cost_u:.4f}" if cost_u else "",
                    "Actual Qty":       _fmt_qty(actl_qty),
                    "Theoretical Qty":  _fmt_qty(theo_qty),
                    "Efficiency %":     f"{abs(effcy):.0%}" if theo_qty else "0%",
                    "Actual Spend ($)": f"${actl_usd:,.2f}" if actl_usd else "$0",
                }

            else:
                ing   = avt["ingredients"].get(item_lower, {})
                sc    = stock.get(item_lower, {})

                end_qty  = ing.get("end_qty",  0.0)
                cost_u   = sc.get("cost", ing.get("unit_cost", 0.0))
                actl_qty = ing.get("actl_qty", 0.0)
                theo_qty = ing.get("theo_qty", 0.0)
                effcy    = ing.get("effcy",    0.0)
                actl_usd = ing.get("actl_usd", 0.0)
                uom      = ing.get("uom", "")

                values = {
                    "End Qty":          _fmt_qty(end_qty),
                    "Cost / Unit ($)":  f"{cost_u:.4f}" if cost_u else "",
                    "Actual Qty":       _fmt_qty(actl_qty),
                    "Theoretical Qty":  _fmt_qty(theo_qty),
                    "Efficiency %":     f"{abs(effcy):.0%}" if effcy else "0%",
                    "Actual Spend ($)": f"${actl_usd:,.2f}" if actl_usd else "$0",
                }

            for metric in _DEEP_DIVE_METRICS:
                records.append({
                    "location":   loc_label,
                    "item":       item_name,
                    "uom":        uom,
                    "metric":     metric,
                    "week_label": week_label,
                    "value":      values[metric],
                })

    return records

# ── Google Sheet writer ────────────────────────────────────────────────────────

def _sheets_api(gc: gspread.Client):
    """Return the raw Sheets API v4 service from the gspread client."""
    return gc.auth  # gspread exposes the underlying credentials


def _batch_update(sh: gspread.Spreadsheet, requests: list[dict]) -> None:
    if not requests:
        return
    sh.batch_update({"requests": requests})


def _color_req(ws_id: int, row: int, col_start: int, col_end: int, color: dict) -> dict:
    return {
        "repeatCell": {
            "range": {
                "sheetId": ws_id,
                "startRowIndex": row, "endRowIndex": row + 1,
                "startColumnIndex": col_start, "endColumnIndex": col_end,
            },
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    }


def _bold_req(ws_id: int, row: int, bold: bool = True) -> dict:
    return {
        "repeatCell": {
            "range": {
                "sheetId": ws_id,
                "startRowIndex": row, "endRowIndex": row + 1,
                "startColumnIndex": 0, "endColumnIndex": 20,
            },
            "cell": {"userEnteredFormat": {"textFormat": {"bold": bold}}},
            "fields": "userEnteredFormat.textFormat.bold",
        }
    }


def write_weekly_data_tab(
    sh: gspread.Spreadsheet,
    rows: list[list],
) -> None:
    """Write (or overwrite) the Weekly Data tab."""
    existing = {ws.title: ws for ws in sh.worksheets()}

    if "Weekly Data" in existing:
        ws = existing["Weekly Data"]
        ws.clear()
    else:
        ws = sh.add_worksheet(title="Weekly Data", rows=500, cols=20)

    # Pad rows to equal length
    max_cols = max((len(r) for r in rows if r), default=1)
    padded   = [r + [""] * (max_cols - len(r)) for r in rows]

    ws.update(padded, "A1")

    # ── Formatting ────────────────────────────────────────────────────────────
    ws_id    = ws.id
    requests = []

    # Clear all existing formatting first so stale formatting doesn't persist
    requests.append({
        "updateCells": {
            "range": {"sheetId": ws_id},
            "fields": "userEnteredFormat",
        }
    })

    for ri, row in enumerate(rows):
        if not row:
            continue
        first = str(row[0]) if row else ""

        if first.startswith("Santo Taco —"):         # title (avoid matching "Santo Taco Water Still")
            requests += [_bold_req(ws_id, ri), _color_req(ws_id, ri, 0, max_cols, MINT_GREEN)]
        elif first.startswith("SLIDE"):             # section header
            requests += [_bold_req(ws_id, ri), _color_req(ws_id, ri, 0, max_cols, DARK_GREEN)]
        elif first in ("Metric", "Taco", "Item"):   # column header
            requests += [_bold_req(ws_id, ri), _color_req(ws_id, ri, 0, max_cols, LIGHT_GREY)]

    # Freeze row 1
    requests.append({
        "updateSheetProperties": {
            "properties": {"sheetId": ws_id, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount",
        }
    })

    _batch_update(sh, requests)
    print(f"  [ok]   Weekly Data tab written ({len(rows)} rows)")


def write_deep_dive_tab(
    sh:       gspread.Spreadsheet,
    records:  list[dict],
    end_date: date,
) -> None:
    """
    Append a new week column to the COGS Deep Dive tab.
    Layout: Location | Item | UOM | Metric | Wk YYYY-MM-DD | Wk YYYY-MM-DD | ...

    On first run the tab is created from scratch.
    On subsequent runs the script reads existing data, adds the new column,
    and rewrites the sheet.
    """
    week_label = _dd_week_col_label(end_date)
    existing   = {ws.title: ws for ws in sh.worksheets()}

    # ── Load or initialise ────────────────────────────────────────────────────
    if "COGS Deep Dive" in existing:
        ws         = existing["COGS Deep Dive"]
        raw        = ws.get_all_values()
        if raw:
            header    = raw[0]
            data_rows = raw[1:]
        else:
            header    = ["Location", "Item", "UOM", "Metric"]
            data_rows = []
    else:
        ws        = sh.add_worksheet(title="COGS Deep Dive", rows=2000, cols=60)
        header    = ["Location", "Item", "UOM", "Metric"]
        data_rows = []

    # ── Add or overwrite the week column ───────────────────────────────────────
    if week_label in header:
        week_col_idx = header.index(week_label)
        print(f"  [update] COGS Deep Dive overwriting column '{week_label}'")
    else:
        week_col_idx = len(header)
        header.append(week_label)

    # Purge stale rows whose item name is no longer in _DEEP_DIVE_ITEMS
    valid_item_names = {item["name"] for item in _DEEP_DIVE_ITEMS}
    data_rows = [
        row for row in data_rows
        if len(row) < 2 or row[1] in valid_item_names
    ]

    # Build a lookup from existing rows: (location, item, uom, metric) → row index
    row_lookup: dict[tuple, int] = {}
    for ri, row in enumerate(data_rows):
        if len(row) >= 4:
            key = (row[0], row[1], row[2], row[3])
            row_lookup[key] = ri

    # Extend existing rows with empty value for new week column
    for row in data_rows:
        while len(row) < len(header):
            row.append("")

    # Upsert records for the new week
    new_keys: set[tuple] = set()
    for rec in records:
        key = (rec["location"], rec["item"], rec["uom"], rec["metric"])
        new_keys.add(key)

        if key in row_lookup:
            ri = row_lookup[key]
            while len(data_rows[ri]) < len(header):
                data_rows[ri].append("")
            data_rows[ri][week_col_idx] = rec["value"]
        else:
            new_row = [""] * len(header)
            new_row[0] = rec["location"]
            new_row[1] = rec["item"]
            new_row[2] = rec["uom"]
            new_row[3] = rec["metric"]
            new_row[week_col_idx] = rec["value"]
            data_rows.append(new_row)
            row_lookup[key] = len(data_rows) - 1

    # ── Sort: Location → Item → Metric order ──────────────────────────────────
    metric_order = {m: i for i, m in enumerate(_DEEP_DIVE_METRICS)}
    loc_order    = {"SoHo": 0, "Union Square": 1}

    data_rows.sort(key=lambda r: (
        loc_order.get(r[0], 99),
        r[1].lower() if r[1] else "",
        metric_order.get(r[3], 99),
    ))

    # ── Write ─────────────────────────────────────────────────────────────────
    all_rows = [header] + data_rows
    ws.clear()
    ws.update(all_rows, "A1")

    # ── Formatting ────────────────────────────────────────────────────────────
    ws_id    = ws.id
    requests = []

    # Header row
    requests += [
        _bold_req(ws_id, 0),
        _color_req(ws_id, 0, 0, len(header), MINT_GREEN),
    ]

    # Freeze first row + first 4 columns
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": ws_id,
                "gridProperties": {
                    "frozenRowCount": 1,
                    "frozenColumnCount": _DD_LEFT_COLS,
                },
            },
            "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
        }
    })

    # Alternate location shading + bold metric headers
    prev_loc  = None
    shade_idx = 0
    shades    = [WHITE, LIGHT_GREY]

    for ri, row in enumerate(data_rows):
        if not row:
            continue
        loc    = row[0] if row else ""
        metric = row[3] if len(row) > 3 else ""

        if loc != prev_loc:
            shade_idx = (shade_idx + 1) % 2
            prev_loc  = loc

        color = DARK_GREEN if loc == "SoHo" and metric == _DEEP_DIVE_METRICS[0] else shades[shade_idx]

        # Bold the "End Qty" row (first metric for each ingredient — visual separator)
        if metric == _DEEP_DIVE_METRICS[0]:
            requests.append(_bold_req(ws_id, ri + 1))

    _batch_update(sh, requests)
    print(f"  [ok]   COGS Deep Dive tab updated — column '{week_label}' added")

# ── Weekly history ─────────────────────────────────────────────────────────────

def update_cogs_history(
    sh:         gspread.Spreadsheet,
    end_date:   date,
    avt_soho:   dict,
    avt_usq:    dict,
    toast_soho: dict | None,
    toast_usq:  dict | None,
) -> None:
    """Append or overwrite weekly summary rows in the 'Weekly History' tab."""
    HISTORY_COLS = [
        "week_start", "week_end", "location", "net_sales",
        "am_orders", "pm_orders", "total_orders",
        "total_cogs", "cogs_pct",
    ]

    try:
        ws = sh.worksheet("Weekly History")
    except Exception as e:
        if "WorksheetNotFound" not in type(e).__name__:
            raise
        ws = sh.add_worksheet(title="Weekly History", rows=500, cols=len(HISTORY_COLS))

    existing = ws.get_all_values()
    if not existing or existing[0] != HISTORY_COLS:
        ws.clear()
        existing = [HISTORY_COLS]

    data_rows = existing[1:]
    # Build lookup: (week_end, location) -> index in data_rows
    row_index = {}
    for i, row in enumerate(data_rows):
        if len(row) >= 3:
            row_index[(row[1], row[2])] = i

    week_start_str = (end_date - timedelta(days=6)).isoformat()
    week_end_str   = end_date.isoformat()

    for loc_name, avt, toast in [("SoHo", avt_soho, toast_soho), ("Union Square", avt_usq, toast_usq)]:
        net_sales    = toast["net_sales"]    if toast else avt["net_sales"]
        am_orders    = toast["am_orders"]    if toast else None
        pm_orders    = toast["pm_orders"]    if toast else None
        total_orders = toast["total_orders"] if toast else None
        total_cogs   = round(sum(ing["actl_usd"] for ing in avt["ingredients"].values() if ing["actl_usd"] > 0), 2)
        cogs_pct     = round(total_cogs / net_sales, 4) if net_sales else None

        new_row = [
            week_start_str,
            week_end_str,
            loc_name,
            round(net_sales, 2),
            am_orders,
            pm_orders,
            total_orders,
            total_cogs,
            cogs_pct,
        ]

        key = (week_end_str, loc_name)
        if key in row_index:
            data_rows[row_index[key]] = new_row
        else:
            data_rows.append(new_row)

    data_rows.sort(key=lambda r: (r[0], r[2]) if len(r) >= 3 else ("", ""))
    ws.update([HISTORY_COLS] + data_rows, "A1")


# ── Sheet initialisation ───────────────────────────────────────────────────────

def get_or_create_sheet(gc: gspread.Client, end_date: date) -> gspread.Spreadsheet:
    """
    Open the sheet by COGS_SHEET_ID if set, otherwise create a new one
    and print its ID for you to save to .env.
    """
    if COGS_SHEET_ID:
        sh = gc.open_by_key(COGS_SHEET_ID)
        print(f"  [ok]   Opened sheet: {sh.url}")
        return sh

    title = f"Santo Taco COGS Deep Dive"
    sh    = gc.create(title)
    sh.share(SHARE_WITH_EMAIL, perm_type="user", role="writer")
    print(f"\n  [NEW]  Created Google Sheet: {sh.url}")
    print(f"         Add to .env:  COGS_SHEET_ID={sh.id}\n")
    return sh

# ── Main ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Santo Taco COGS Deep Dive")
    p.add_argument(
        "--week-ending",
        metavar="YYYY-MM-DD",
        help="Sunday date of the week to process (default: most recent Sunday)",
    )
    p.add_argument(
        "--local-avt-soho",
        metavar="PATH",
        help="Path to local AVT SoHo Excel file (skip Gmail fetch)",
    )
    p.add_argument(
        "--local-avt-usq",
        metavar="PATH",
        help="Path to local AVT Union Square Excel file (skip Gmail fetch)",
    )
    p.add_argument(
        "--local-stock",
        metavar="PATH",
        help="Path to local Stock Count Detail Excel file (skip Gmail fetch)",
    )
    return p.parse_args()


def main() -> None:
    args     = parse_args()
    run_date = date.today()

    if args.week_ending:
        try:
            end_date = date.fromisoformat(args.week_ending)
        except ValueError:
            sys.exit(f"[ERROR] --week-ending must be YYYY-MM-DD, got: {args.week_ending}")
        if end_date.weekday() != 6:
            sys.exit(
                f"[ERROR] --week-ending must be a Sunday; "
                f"{end_date.isoformat()} is a {end_date.strftime('%A')}."
            )
    else:
        end_date = most_recent_sunday(run_date)

    begin_date = end_date - timedelta(days=6)

    print()
    print("Santo Taco COGS Deep Dive".center(72))
    print(f"Run date  : {run_date.isoformat()}".center(72))
    print(f"Period    : {begin_date.isoformat()} → {end_date.isoformat()}".center(72))
    print()


    # ── Step 1: Fetch / load Excel files ──────────────────────────────────────
    print("[1/5] Loading R365 source files…")

    def _load(local_path, subject, label):
        if local_path:
            print(f"  [local] {label}: {local_path}")
            with open(local_path, "rb") as f:
                return f.read()
        print(f"  [gmail] Fetching '{subject}'…")
        return fetch_excel_attachment(subject)

    if args.local_avt_soho or args.local_avt_usq:
        raw_avt_soho = _load(args.local_avt_soho, AVT_SUBJECT, "AVT SoHo")
        raw_avt_usq  = _load(args.local_avt_usq,  AVT_SUBJECT, "AVT USQ")
    else:
        print(f"  [gmail] Fetching AVT emails (subject: '{AVT_SUBJECT}')…")
        raw_avt_soho, raw_avt_usq = fetch_avt_attachments(AVT_SUBJECT)
    raw_stock = _load(args.local_stock, STOCK_SUBJECT, "Stock Count")

    missing = []
    if not raw_avt_soho: missing.append("AVT SoHo")
    if not raw_avt_usq:  missing.append("AVT Union Square")
    if not raw_stock:    missing.append("Stock Count Detail")
    if missing:
        sys.exit(f"\n[ERROR] Could not load: {', '.join(missing)}\n")

    # ── Step 2: Parse R365 files ───────────────────────────────────────────────
    print("\n[2/5] Parsing R365 files…")
    avt_soho = parse_avt(raw_avt_soho, sheet_index=0)
    avt_usq  = parse_avt(raw_avt_usq,  sheet_index=0)
    stock    = parse_stock_count(raw_stock)

    print(f"  AVT SoHo   : {len(avt_soho['ingredients'])} ingredients | "
          f"net sales ${avt_soho['net_sales']:,.2f} | "
          f"{avt_soho['begin_date']} → {avt_soho['end_date']}")
    print(f"  AVT USQ    : {len(avt_usq['ingredients'])} ingredients | "
          f"net sales ${avt_usq['net_sales']:,.2f} | "
          f"{avt_usq['begin_date']} → {avt_usq['end_date']}")
    print(f"  Stock SoHo : {len(stock['soho'])} items")
    print(f"  Stock USQ  : {len(stock['usq'])} items")

    # ── Step 3: Fetch Toast sales + orders ────────────────────────────────────
    print("\n[3/5] Fetching sales from Toast…")
    toast_soho = None
    toast_usq  = None

    if not TOAST_CLIENT_ID or not TOAST_CLIENT_SECRET:
        print("  [warn] TOAST credentials not set — skipping Toast fetch")
    else:
        try:
            token = toast_authenticate()

            print(f"  › SoHo…", end=" ", flush=True)
            toast_soho = toast_fetch_sales_and_orders(
                token, TOAST_GUIDS["soho"], begin_date, end_date
            )
            print(f"${toast_soho['net_sales']:,.2f} | "
                  f"{toast_soho['am_orders']} Lunch / {toast_soho['pm_orders']} Dinner orders")

            print(f"  › Union Square…", end=" ", flush=True)
            toast_usq = toast_fetch_sales_and_orders(
                token, TOAST_GUIDS["usq"], begin_date, end_date
            )
            print(f"${toast_usq['net_sales']:,.2f} | "
                  f"{toast_usq['am_orders']} Lunch / {toast_usq['pm_orders']} Dinner orders")

        except Exception as exc:
            print(f"  [warn] Toast fetch failed: {exc} — slide 3 will use R365 sales")

    # ── Override hours with Places API (more accurate than first/last order) ───
    for loc, toast_data in [("soho", toast_soho), ("usq", toast_usq)]:
        if toast_data is None:
            continue
        am_h, pm_h = fetch_open_hours(PLACE_IDS[loc], begin_date, end_date)
        if am_h > 0 or pm_h > 0:
            toast_data["am_hours"] = am_h
            toast_data["pm_hours"] = pm_h

    # ── Step 4: Build output data ──────────────────────────────────────────────
    print("\n[4/5] Assembling sheet data…")
    weekly_rows = build_weekly_data(
        avt_soho, avt_usq, end_date,
        toast_soho=toast_soho,
        toast_usq=toast_usq,
    )
    deep_records = build_deep_dive_rows(
        avt_soho, avt_usq,
        stock["soho"], stock["usq"],
        end_date,
    )
    print(f"  Weekly Data : {len(weekly_rows)} rows")
    print(f"  Deep Dive   : {len(deep_records)} records "
          f"({len(deep_records) // len(_DEEP_DIVE_METRICS)} ingredients)")

    # ── Step 5: Write Google Sheet ─────────────────────────────────────────────
    print("\n[5/5] Writing to Google Sheet…")
    gc = _get_sheets_client()
    sh = get_or_create_sheet(gc, end_date)

    write_weekly_data_tab(sh, weekly_rows)
    write_deep_dive_tab(sh, deep_records, end_date)
    update_cogs_history(sh, end_date, avt_soho, avt_usq, toast_soho, toast_usq)
    print("  [ok]   Weekly History updated")

    print()
    print("=" * 72)
    print(f"  Sheet: {sh.url}")
    print("=" * 72)
    print()

if __name__ == "__main__":
    main()
